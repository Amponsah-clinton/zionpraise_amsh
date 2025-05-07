from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.contrib.auth.models import User
from django.contrib import messages
from .forms import StudentRegistrationForm

def add_student(request):
    """Handles student registration and auto-creates a parent account."""
    if request.method == "POST":
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)  # Delay saving
            student.save()  # Triggers save() method to auto-create parent
            user = authenticate(username=student.user.username, password=request.POST.get('password1'))
            if user:
                login(request, user)
            messages.success(request, "Student registered successfully!")
            return redirect('add_student')  # Redirect to clear form and avoid duplicate submissions
    else:
        form = StudentRegistrationForm()

    return render(request, 'add_student.html', {'form': form})


from django.db import models
from django.shortcuts import render, redirect
from .forms import ParentLoginForm
from django.shortcuts import render, redirect
from .forms import ParentLoginForm
from .models import Student

def parent_login(request):
    if request.method == "POST":
        form = ParentLoginForm(request.POST)
        if form.is_valid():
            students = form.cleaned_data["students"]  # Now works because it's in cleaned_data

            # Ensure that there's at least one student
            if not students:
                form.add_error('username', "No valid student found with the given parent details.")
                return render(request, 'parent_login.html', {'form': form})

            # Store session details
            request.session["parent_logged_in"] = True
            request.session["parent_identifier"] = request.POST.get("username")
            request.session["student_ids"] = [student.id for student in students]

            # Redirect to the first student's detail page
            return redirect("student_detail", student_id=students[0].id)
        else:
            print("❌ Form errors:", form.errors)
    else:
        form = ParentLoginForm()

    return render(request, "parent_login.html", {"form": form})




from django.contrib.auth.decorators import login_required
from django.utils.timezone import now
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, F, Window
from django.db.models.functions import Rank
from collections import defaultdict
from .models import Student, Score, Subject, AcademicCalendar, Term
from django.db.models import Q  # Import Q directly

def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    subjects = Subject.objects.all()

    selected_term_id = request.GET.get("term")
    selected_term = None
    if selected_term_id:
        try:
            selected_term = Term.objects.get(id=selected_term_id)
        except Term.DoesNotExist:
            selected_term = None

    assessments = Score.objects.filter(student=student)
    if selected_term:
        assessments = assessments.filter(term=selected_term)

    assessment_choices = [
        ("class_test", "Class Test"),
        ("monthly_exam", "Monthly Exam"),
        ("midterm_exam", "Midterm Exam"),
        ("appraisal_exam", "Appraisal Exam"),
        ("end_of_term_exam", "End of Term Exam"),
    ]

    subject_scores = defaultdict(dict)
    total_scores = {key: 0 for key, _ in assessment_choices}

    for assessment in assessments:
        subject_name = assessment.subject.name
        assessment_type = assessment.assessment_type
        score = assessment.score

        subject_scores[subject_name][assessment_type] = score
        total_scores[assessment_type] += score

    for subject in subjects:
        for key, _ in assessment_choices:
            subject_scores[subject.name].setdefault(key, 0)

    # Calculate the rank for each assessment type separately
    rank_per_assessment = {}
    for key, _ in assessment_choices:
        # Filter scores for the selected term and grade for this assessment type
        score_query = Score.objects.filter(student__grade=student.grade, assessment_type=key)
        if selected_term:
            score_query = score_query.filter(term=selected_term)

        # Rank students based on the score for this assessment type
        ranked_students = (
            score_query.values("student_id", "score")
            .annotate(total_score=Sum("score"))
            .annotate(rank=Window(expression=Rank(), order_by=F("total_score").desc()))
        )

        # Find the rank for this specific student
        student_rank = next(
            (entry["rank"] for entry in ranked_students if entry["student_id"] == student.id), "-"
        )
        rank_per_assessment[key] = student_rank

    total_students = Student.objects.filter(grade=student.grade).count()
    academic_calendar = AcademicCalendar.objects.filter(date__gte=now()).order_by("date")
    all_terms = Term.objects.all().order_by("-start_date")
    column_count = len(assessment_choices) + 1

    # Add after fetching the main student
    parent_identifier = request.session.get("parent_identifier")
    siblings = Student.objects.filter(
        Q(parent_email=parent_identifier) | Q(parent_phone=parent_identifier)
    ).order_by("name")
    show_dashboard_button = False
    if request.user.is_authenticated:
        show_dashboard_button = (
            request.user.is_superuser or
            Teacher.objects.filter(user=request.user).exists()
        )

    context = {
        "student": student,
        "show_dashboard_button": show_dashboard_button,
        "siblings": siblings,
        "subjects": subjects,
        "assessment_choices": assessment_choices,
        "subject_scores": dict(subject_scores),
        "total_scores": total_scores,
        "rank_per_assessment": rank_per_assessment,
        "total_students": total_students,
        "academic_calendar": academic_calendar,
        "column_count": column_count,
        "selected_term": selected_term,
        "terms": all_terms,
    }

    return render(request, "student_detail.html", context)


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import Sum

from .models import Student, Score, Subject

def download_student_scores(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    subjects = Subject.objects.all()
    assessments = Score.objects.filter(student=student)

    assessment_choices = [
        ("class_test", "Class Test"),
        ("monthly_exam", "Monthly Exam"),
        ("midterm_exam", "Midterm Exam"),
        ("appraisal_exam", "Appraisal Exam"),
        ("end_of_term_exam", "End of Term Exam"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Report Card"

    # === Styles ===
    center = Alignment(horizontal='center')
    bold = Font(name='Times New Roman', bold=True)
    larger_bold = Font(name='Times New Roman', bold=True, size=13)
    regular = Font(name='Times New Roman')
    total_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # light yellow
    rank_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")   # light red

    # === Student Info Header ===
    details = [
        f"Name: {student.name}",
        f"Grade: {student.grade.upper()}",
        f"Parent: {student.parent_name or 'None'}",
        f"Parent Email: {student.parent_email or 'None'}",
        f"Parent Phone: {student.parent_phone or 'None'}"
    ]

    for idx, line in enumerate(details, 1):
        cell = ws.cell(row=idx, column=1, value=line)
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=6)
        cell.alignment = center
        cell.font = bold

    start_row = len(details) + 2

    # === Header Row (no Rank column) ===
    header = ['Subject'] + [label for _, label in assessment_choices]
    for col_num, col_name in enumerate(header, 1):
        cell = ws.cell(row=start_row, column=col_num, value=col_name)
        cell.font = bold
        cell.alignment = center

    # === Scores Table ===
    subject_scores = []
    total_scores_per_assessment = [0] * len(assessment_choices)

    for subject in subjects:
        row = [subject.name]
        for idx, (atype, _) in enumerate(assessment_choices):
            score = assessments.filter(subject=subject, assessment_type=atype).aggregate(score=Sum('score'))['score'] or 0
            total_scores_per_assessment[idx] += score
            row.append(score)
        subject_scores.append(row)

    for i, row_data in enumerate(subject_scores, start=start_row + 1):
        for j, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.font = regular
            cell.alignment = center

    # === Total Row ===
    total_row = start_row + len(subject_scores) + 1
    ws.cell(row=total_row, column=1, value="Total").font = larger_bold
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=1).fill = total_fill

    for idx, total in enumerate(total_scores_per_assessment):
        cell = ws.cell(row=total_row, column=idx + 2, value=total)
        cell.font = larger_bold
        cell.alignment = center
        cell.fill = total_fill

    # === Rank Row ===
    rank_row = total_row + 1
    ws.cell(row=rank_row, column=1, value="📊 Rank").font = bold
    ws.cell(row=rank_row, column=1).alignment = center
    ws.cell(row=rank_row, column=1).fill = rank_fill

    all_students = Student.objects.all()
    for idx, (atype, _) in enumerate(assessment_choices):
        student_total = sum(
            Score.objects.filter(student=student, subject=subject, assessment_type=atype)
            .aggregate(score=Sum('score'))['score'] or 0
            for subject in subjects
        )

        if student_total == 0:
            rank = "N/A"
        else:
            student_totals = []
            for stu in all_students:
                total = sum(
                    Score.objects.filter(student=stu, subject=subject, assessment_type=atype)
                    .aggregate(score=Sum('score'))['score'] or 0
                    for subject in subjects
                )
                student_totals.append((stu.id, total))

            student_totals.sort(key=lambda x: x[1], reverse=True)
            rank = next((i + 1 for i, (sid, _) in enumerate(student_totals) if sid == student.id), "N/A")

        cell = ws.cell(row=rank_row, column=idx + 2, value=rank)
        cell.font = bold
        cell.alignment = center
        cell.fill = rank_fill

    # === Adjust Column Widths ===
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # === Set Default Font ===
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Times New Roman')

    # === Response with File Named after Student ===
    safe_name = student.name.replace(" ", "_")
    filename = f"student_{safe_name}_report_card.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response





from app.models import Teacher
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from .models import Teacher, Subject
from .forms import TeacherRegistrationForm, ScoreForm

from django.shortcuts import render, redirect
from .forms import TeacherRegistrationForm
from .models import Teacher, Subject
from django.contrib.auth.models import User

def teacher_register(request):
    if request.method == "POST":
        form = TeacherRegistrationForm(request.POST)
        if form.is_valid():
            teacher = form.save(commit=False)
            subject_id = request.POST.get('subject')

            # Check subject is already assigned
            if subject_id:
                subject = Subject.objects.get(id=subject_id)
                if hasattr(subject, 'teacher'):
                    form.add_error('subject', 'This subject is already assigned to another teacher.')
                    return render(request, 'teacher_register.html', {'form': form})

                teacher.subject = subject

            teacher.save()
            return redirect('teacher_dashboard')  # Replace with your actual route
    else:
        form = TeacherRegistrationForm()

    return render(request, 'teacher_register.html', {'form': form})



from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Teacher, Student, Score
from django.db import transaction
from django.contrib import messages

ASSESSMENT_CHOICES = [
    ('class_test', 'Class Test'),
    ('monthly_exam', 'Monthly Exam'),
    ('midterm_exam', 'Midterm Exam'),
    ('appraisal_exam', 'Appraisal Exam'),
    ('end_of_term_exam', 'End of Term Exam'),
]




from django.contrib import messages
from django.shortcuts import render, redirect
from django.db import transaction
from django.contrib.auth.decorators import login_required
from .models import Teacher, Student, Score



from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import TeacherLoginForm
from .models import Teacher

def teacher_login(request):
    if request.method == 'POST':
        form = TeacherLoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            try:
                teacher = Teacher.objects.get(email=email)

                # Compare plain passwords
                if teacher.password == password:
                    user = teacher.user
                    login(request, user)  # Still logs in the associated Django User
                    request.session['teacher_id'] = teacher.id
                    return redirect('teacher_dashboard')
                else:
                    messages.error(request, "Invalid password.")
            except Teacher.DoesNotExist:
                messages.error(request, "Teacher with this email does not exist.")
    else:
        form = TeacherLoginForm()

    return render(request, 'teacher_login.html', {'form': form})





from openpyxl.worksheet.datavalidation import DataValidation
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import Term, Student, Score

@login_required(login_url='teacher_login')
def export_student_scores(request):
    term_id = request.GET.get("term_id")
    assessment_type = request.GET.get("assessment_type")
    selected_grade = request.GET.get("grade")

    if not term_id or not assessment_type or not selected_grade:
        messages.error(request, "Please select a term, assessment type, and grade.")
        return redirect('teacher_dashboard')

    try:
        term = Term.objects.get(id=term_id)
    except Term.DoesNotExist:
        messages.error(request, "Invalid term.")
        return redirect('teacher_dashboard')

    teacher = request.user.teacher_profile
    subject = teacher.subject

    students = Student.objects.filter(grade=selected_grade)

    response = HttpResponse(content_type="application/vnd.ms-excel")
    filename = f"scores_{selected_grade}_{term.name}_{assessment_type}_{subject.name}.xlsx".replace(" ", "_")
    response["Content-Disposition"] = f"attachment; filename={filename}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Scores"

    # Header info
    ws['A1'] = f"Subject: {subject.name}"
    ws['B1'] = f"Teacher: {teacher.name}"
    ws['C1'] = f"Term: {term.name}"
    ws['D1'] = f"Assessment Type: {assessment_type}"
    ws['E1'] = f"Grade: {selected_grade}"

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}1'].font = Font(size=12, bold=True, color="FF0000")
        ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 39.5 
    ws.row_dimensions[1].height = 30

    # Add headings
    headings = ["Student ID", "Name", "Score"]
    ws.append(headings)

    for col_num, heading in enumerate(headings, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = heading
        cell.font = Font(size=12, bold=True, color="000000")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = 35

    # Add validation to Score column (C)
    score_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1=0,
        formula2=100,
        showErrorMessage=True,
        errorTitle="Invalid Score",
        error="Please enter a number between 0 and 100.",
    )
    ws.add_data_validation(score_validation)

    row_index = 3
    for student in students:
        score_obj = Score.objects.filter(
            student=student,
            term=term,
            assessment_type=assessment_type,
            subject=subject
        ).first()

        score_value = (
            float(score_obj.score)
            if score_obj and isinstance(score_obj.score, (int, float)) and 0 <= score_obj.score <= 100
            else None
        )

        ws.cell(row=row_index, column=1, value=student.student_id)
        ws.cell(row=row_index, column=2, value=student.name)

        score_cell = ws.cell(row=row_index, column=3)
        if score_value is not None:
            score_cell.value = round(score_value, 2)
        else:
            score_cell.value = ""

        score_cell.number_format = '0.00'  # 👈 format for display and manual entry
        score_validation.add(f"C{row_index}")
        row_index += 1

    # 👇 Format entire remaining column C with 0.00 to cover future entries by user
    for i in range(row_index, row_index + 100):  # Future rows for new data entry
        cell = ws.cell(row=i, column=3)
        cell.number_format = '0.00'
        score_validation.add(f"C{i}")

    wb.save(response)
    return response




import csv
import openpyxl
from io import TextIOWrapper
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from .models import Score, Student, Term, Subject, Teacher

ASSESSMENT_CHOICES = [
    ('class_test', 'Class Test'),
    ('monthly_exam', 'Monthly Exam'),
    ('midterm_exam', 'Midterm Exam'),
    ('appraisal_exam', 'Appraisal Exam'),
    ('end_of_term_exam', 'End of Term Exam'),
]


def extract_metadata_from_row(row):
    meta_dict = {}
    for cell in row:
        if cell and ":" in str(cell):
            key, value = str(cell).split(":", 1)
            meta_dict[key.strip().lower()] = value.strip()
    return meta_dict


from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.timezone import now
from io import TextIOWrapper
import csv
import openpyxl

from .models import Teacher, Term, Subject, Student, Score
from .models import ASSESSMENT_CHOICES


@login_required
def upload_student_scores(request):
    teacher = None

    # Check if user is staff (admin)
    if request.user.is_staff:
        is_admin = True
    else:
        # Check if session has a teacher
        teacher_id = request.session.get("teacher_id")
        if not teacher_id:
            messages.error(request, "You must be logged in as a teacher or admin.")
            return redirect("teacher_login")

        teacher = Teacher.objects.filter(id=teacher_id).first()
        if not teacher:
            messages.error(request, "Teacher not found.")
            return redirect("teacher_login")

        is_admin = False

    terms = Term.objects.order_by("-start_date")

    if request.method == "POST":
        file = request.FILES.get("csv_file")

        if not file:
            messages.error(request, "Please upload a file.")
            return redirect("upload_student_scores")

        file_name = file.name.lower()
        created, skipped = 0, 0

        try:
            if file_name.endswith(".csv"):
                csv_data = TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.reader(csv_data)
                rows = list(reader)
                metadata_row = rows[0]
                headers = rows[1]
                data_rows = rows[2:]
            elif file_name.endswith((".xls", ".xlsx")):
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                metadata_row = [cell.value for cell in sheet[1]]
                headers = [cell.value for cell in sheet[2]]
                data_rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=3)]
            else:
                messages.error(request, "Unsupported file format. Use .csv, .xls, or .xlsx.")
                return redirect("upload_student_scores")
        except Exception as e:
            messages.error(request, f"Error processing file: {e}")
            return redirect("upload_student_scores")

        meta_dict = extract_metadata_from_row(metadata_row)

        subject_name = meta_dict.get("subject")
        term_name = meta_dict.get("term")
        assessment_type = meta_dict.get("assessment type")
        grade_name = meta_dict.get("grade")

        if not subject_name or not term_name or not assessment_type or not grade_name:
            messages.error(request, "Missing metadata: Subject, Term, Assessment Type, or Grade.")
            return redirect("upload_student_scores")

        term = Term.objects.filter(name__iexact=term_name).first()
        subject = Subject.objects.filter(name__iexact=subject_name).first()

        if not term:
            messages.error(request, f"Invalid term: {term_name}")
            return redirect("upload_student_scores")
        if not subject:
            messages.error(request, f"Invalid subject: {subject_name}")
            return redirect("upload_student_scores")

        for row in data_rows:
            row_dict = dict(zip(headers, row))
            student_id = row_dict.get("Student ID")
            score_value = row_dict.get("Score")

            if not student_id or score_value in (None, ""):
                skipped += 1
                continue

            student = Student.objects.filter(student_id=student_id, grade__iexact=grade_name).first()
            if not student:
                skipped += 1
                continue

            if Score.objects.filter(
                student=student,
                term=term,
                subject=subject,
                assessment_type=assessment_type
            ).exists():
                skipped += 1
                continue

            try:
                score_float = float(score_value)
            except (ValueError, TypeError):
                skipped += 1
                continue

            Score.objects.create(
                student=student,
                subject=subject,
                assessment_type=assessment_type,
                term=term,
                score=score_float,
                date=now().date(),
            )
            created += 1

        messages.success(request, f"✅ Uploaded: {created}, Skipped: {skipped}")
        return redirect("upload_student_scores")

    return render(request, "upload_scores.html", {
        "terms": terms,
        "assessment_choices": ASSESSMENT_CHOICES
    })





from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import Teacher, Student, Score, AcademicCalendar, Term
from django.contrib.auth.decorators import login_required
from django.utils.timezone import now

@login_required(login_url='teacher_login')
def teacher_dashboard(request):
    teacher_id = request.session.get('teacher_id')
    if not teacher_id:
        messages.error(request, "Session expired. Please log in again.")
        return redirect('teacher_login')

    try:
        teacher = Teacher.objects.get(id=teacher_id)
    except Teacher.DoesNotExist:
        messages.error(request, "Invalid teacher account.")
        return redirect('teacher_login')

    grades = Student.objects.values_list("grade", flat=True).distinct()
    students = []
    selected_grade = request.GET.get("grade")
    selected_assessment = request.GET.get("assessment_type")

    if selected_grade and selected_assessment:
        students = Student.objects.filter(grade=selected_grade)

    if request.method == "POST":
        selected_grade = request.POST.get("grade")
        selected_assessment = request.POST.get("assessment_type")

        student_ids = request.POST.getlist("student_id")
        scores = request.POST.getlist("score")

        # Get the current date or the date provided in the POST request
        score_date = now().date()

        # Fetch the term based on the current date
        term = Term.objects.filter(start_date__lte=score_date, end_date__gte=score_date).first()

        if term is None:
            messages.error(request, "No active term found for the current date.")
            return redirect('teacher_dashboard')

        existing_scores = Score.objects.filter(
            student_id__in=student_ids,
            assessment_type=selected_assessment,
            subject=teacher.subject
        ).values_list("student_id", flat=True)

        with transaction.atomic():
            for student_id, score in zip(student_ids, scores):
                if score.strip():
                    if int(student_id) in existing_scores:
                        messages.warning(request, f"Score for student ID {student_id} already exists under this assessment and subject.")
                    else:
                        # Create the score entry with the assigned term
                        Score.objects.create(
                            student_id=student_id,
                            assessment_type=selected_assessment,
                            subject=teacher.subject,
                            score=int(score),
                            date=score_date,  # Optional: specify the date if needed
                            term=term  # Automatically assign the term
                        )
                        messages.success(request, "Scores saved successfully!")

        return redirect('teacher_dashboard')

    academic_calendar = AcademicCalendar.objects.all()
    terms = Term.objects.all().order_by("start_date")
    context = {
        "teacher": teacher,
        "grades": grades,
        "students": students,
        "assessment_choices": ASSESSMENT_CHOICES,
        "selected_grade": selected_grade,
        "selected_assessment": selected_assessment,
        "academic_calendar": academic_calendar,  
        'terms': terms,
    }
    return render(request, "teacher_dashboard.html", context)


def teacher_logout(request):
    logout(request)
    request.session.flush()  
    return redirect('teacher_login')

import csv
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Score, Student, Teacher

def export_scores_csv(request):
    """Export student scores as CSV based on selected grade, assessment type, and teacher's subject."""
    teacher_id = request.session.get('teacher_id')
    if not teacher_id:
        return HttpResponse("Unauthorized", status=401)

    teacher = get_object_or_404(Teacher, id=teacher_id)

    grade = request.GET.get("grade")
    assessment_type = request.GET.get("assessment_type")

    if not grade or not assessment_type:
        return HttpResponse("Grade and assessment type are required!", status=400)

    scores = Score.objects.filter(
        student__grade=grade,
        assessment_type=assessment_type,
        subject=teacher.subject  
    ).select_related("student").order_by("-score")  

    if not scores.exists():
        return HttpResponse("No scores available for the selected grade and assessment.", status=404)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="scores_{teacher.subject}_{grade}_{assessment_type}.csv"'

    writer = csv.writer(response)
    writer.writerow(["Rank", "Student ID", "Name", "Grade", "Assessment Type", "Subject", "Score"])

    for rank, score in enumerate(scores, start=1):
        writer.writerow([
            rank, 
            score.student.student_id, 
            score.student.name, 
            score.student.grade, 
            score.assessment_type, 
            score.subject, 
            score.score
        ])

    return response


from django.http import JsonResponse
from .models import Student

def student_list(request):
    selected_grade = request.GET.get("grade")
    search_query = request.GET.get("search")

    grades = Student.objects.values_list("grade", flat=True).distinct()

    students = Student.objects.all()
    if selected_grade:
        students = students.filter(grade=selected_grade)
    if search_query:
        students = students.filter(name__icontains=search_query)

    # If it's an AJAX request, return JSON data
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        data = []
        for student in students:
            data.append({
                "name": student.name,
                "grade": student.grade,
                "id": student.id,
            })
        return JsonResponse({"students": data})

    return render(request, "student_list.html", {
        "students": students,
        "grades": grades,
        "selected_grade": selected_grade,
    })



def export_students_csv(request):
    """Exports selected grade students as a CSV file."""
    selected_grade = request.GET.get("grade")
    students = Student.objects.filter(grade=selected_grade) if selected_grade else Student.objects.all()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="students.csv"'

    writer = csv.writer(response)
    writer.writerow(["Student ID", "Name", "Grade", "Parent Email", "Parent Phone"])

    for student in students:
        writer.writerow([student.student_id, student.name, student.grade, student.parent_email, student.parent_phone])

    return response





from django.shortcuts import render
from django.db.models import Sum, Q

def analysis_view(request):
    selected_grade = request.GET.get("grade", "")
    selected_category = request.GET.get("category", "class_test")
    selected_term = request.GET.get("term", "")  # New line to capture selected term

    students = Student.objects.all()

    # Apply grade filter
    if selected_grade:
        if selected_grade == "7":
            students = students.filter(grade__in=["7a", "7b", "7c"])
        elif selected_grade == "8":
            students = students.filter(grade__in=["8a", "8b"])
        else:
            students = students.filter(grade=selected_grade)

    # Apply assessment type and term filter
    filter_condition = Q(scores__assessment_type=selected_category)
    if selected_term:
        filter_condition &= Q(scores__term=selected_term)

    student_scores = students.annotate(
        total_score=Sum("scores__score", filter=filter_condition)
    ).order_by("-total_score")

    for student in student_scores:
        student.total_score = student.total_score or 0  # Avoid None values

    context = {
        "students": student_scores,
        "grades": [
            ("7", "All Grade 7"), ("8", "All Grade 8"), 
            ("7a", "7A"), ("7b", "7B"), ("7c", "7C"),
            ("8a", "8A"), ("8b", "8B"), ("9", "9")
        ],
        "assessment_categories": [
            ('class_test', 'Class Test'),
            ('monthly_exam', 'Monthly Exam'),
            ('midterm_exam', 'Midterm Exam'),
            ('appraisal_exam', 'Appraisal Exam'),
            ('end_of_term_exam', 'End of Term Exam'),
        ],
        "terms": [("1", "Term 1"), ("2", "Term 2"), ("3", "Term 3")],
        "selected_grade": selected_grade,
        "selected_category": selected_category,
        "selected_term": selected_term,
    }

    return render(request, "analysis.html", context)




import csv
from django.http import HttpResponse

def export_to_excel(students, subjects, student_subject_scores, total_scores, subject_ranks, selected_category, selected_grade):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{selected_grade}_{selected_category}_assessment.csv"'

    writer = csv.writer(response)

    # First row: Merging with "Zion Praise Educational Complex"
    writer.writerow(['' for _ in range(len(subjects) + 3)])  # Create an empty row for merging
    writer.writerow(['' for _ in range(2)] + ['ZION PRAISE EDUCATIONAL COMPLEX'] + ['' for _ in range(len(subjects) + 1)])

    # Second row: Adding assessment type and grade
    writer.writerow(['' for _ in range(2)] + [f"Assessment Type: {selected_category.upper()}"] + [f"Grade: {selected_grade}"] + ['' for _ in range(len(subjects) + 1)])

    # Header row
    header = ['   RANK   ', '                  STUDENT NAME                 '] + \
             [f"   {subject.name.upper()}   " for subject in subjects] + ['   TOTAL SCORE   ']
    writer.writerow(header)

    for rank, student in enumerate(students, start=1):
        # Rank column remains as a simple rank
        rank_col = f"{rank}"

        # Increase width for the student name column (column B)
        name_col = f"  {student.name:<50}"  # Increased width (50 characters for the name)

        row = [rank_col, name_col]

        for subject in subjects:
            score = student_subject_scores.get(student.id, {}).get(subject.name)
            subj_rank = subject_ranks.get(subject.name, {}).get(student.id)

            if score is not None and subj_rank:
                score_str = f"{score:.2f} ({subj_rank})"
            elif score is not None:
                score_str = f"{score:.2f}"
            else:
                score_str = ""  # Empty string for missing scores

            row.append(score_str)

        total = total_scores.get(student.id, '')
        row.append(f"{total:.2f}" if isinstance(total, (int, float)) else total)

        writer.writerow(row)

    return response






import csv
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Sum
from .models import Student, Subject, Score

def assessment_details_view(request):
    selected_grade = request.GET.get("grade", "")
    selected_category = request.GET.get("category", "class_test")

    # Filter students
    students = Student.objects.all()
    if selected_grade:
        if selected_grade == "7":
            students = students.filter(grade__in=["7a", "7b", "7c"])
        elif selected_grade == "8":
            students = students.filter(grade__in=["8a", "8b"])
        else:
            students = students.filter(grade=selected_grade)

    subjects = Subject.objects.all()

    # Collect subject-wise scores
    student_subject_scores = {}
    total_scores = {}

    for student in students:
        scores = Score.objects.filter(
            student=student, assessment_type=selected_category
        ).values("subject__name", "score")

        student_scores = {score["subject__name"]: score["score"] for score in scores}
        student_subject_scores[student.id] = student_scores

        total_scores[student.id] = sum(student_scores.values())

    ranked_students = sorted(students, key=lambda s: total_scores.get(s.id, 0), reverse=True)

    subject_ranks = {subject.name: {} for subject in subjects}
    
    for subject in subjects:
        subject_scores = {student.id: student_subject_scores.get(student.id, {}).get(subject.name, 0) for student in students}
        sorted_scores = sorted(subject_scores.items(), key=lambda x: x[1], reverse=True)

        # Assign ranks
        rank = 1
        for i, (student_id, score) in enumerate(sorted_scores):
            if i > 0 and sorted_scores[i - 1][1] > score:
                rank = i + 1
            subject_ranks[subject.name][student_id] = rank

    # Export data if requested
    if request.GET.get("export") == "excel":
        return export_to_excel(ranked_students, subjects, student_subject_scores, total_scores, subject_ranks, selected_category, selected_grade)

    context = {
        "students": ranked_students,
        "subjects": subjects,
        "student_subject_scores": student_subject_scores,
        "total_scores": total_scores,
        "subject_ranks": subject_ranks,
        "selected_grade": selected_grade,
        "selected_category": selected_category,
    }
    return render(request, "assessment_details.html", context)



from django.shortcuts import render
from .models import Student, Score

# Updated Grade Choices including "All Grade 7" and "All Grade 8"
GRADE_CHOICES = [
    ('7a', '7A'), ('7b', '7B'), ('7c', '7C'),
    ('8a', '8A'), ('8b', '8B'),
    ('9', '9'),
    ('all_7', 'All Grade 7'),
    ('all_8', 'All Grade 8')
]

# List of all subjects
SUBJECTS = [
    'Creative Arts', 'Career Tech', 'R.M.E', 'French', 'Twi',
    'Computing', 'English', 'Social Studies', 'Int. Science', 'Mathematics'
]

from django.db.models import Sum

def performance_view(request):
    selected_grade_class = request.GET.get('grade', 'all')
    selected_assessments = request.GET.getlist('assessment')
    selected_term = request.GET.get('term', 'all')

    if selected_grade_class == 'all':
        students = Student.objects.all()
    elif selected_grade_class == 'all_7':
        students = Student.objects.filter(grade__in=['7a', '7b', '7c'])
    elif selected_grade_class == 'all_8':
        students = Student.objects.filter(grade__in=['8a', '8b'])
    else:
        students = Student.objects.filter(grade=selected_grade_class)

    scores_data = []
    for student in students:
        scores = {}
        score_changes = []
        total_change = 0

        for assessment in selected_assessments:
            query = student.scores.filter(assessment_type=assessment)
            if selected_term != 'all':
                query = query.filter(term__id=selected_term)

            # Sum all subject scores under this assessment type for the term(s)
            total_score = query.aggregate(total=Sum('score'))['total']

            if total_score is not None:
                scores[assessment] = total_score
                score_changes.append(total_score)
            else:
                scores[assessment] = 'N/A'

        if len(score_changes) >= 2:
            total_change = score_changes[-1] - score_changes[0]

        scores_data.append({
            'student': student,
            'scores': scores,
            'score_changes': score_changes,
            'total_change': total_change
        })

    most_improved = max(scores_data, key=lambda x: x['total_change'], default=None)
    most_declined = min(scores_data, key=lambda x: x['total_change'], default=None)

    most_improved_msg = f"{most_improved['student'].name} improved by {most_improved['total_change']} points." if most_improved and most_improved['total_change'] > 0 else "No improvement."
    most_declined_msg = f"{most_declined['student'].name} declined by {abs(most_declined['total_change'])} points." if most_declined and most_declined['total_change'] < 0 else "No decline."

    context = {
        'students': scores_data,
        'selected_assessments': selected_assessments,
        'most_improved_msg': most_improved_msg,
        'most_declined_msg': most_declined_msg,
        'selected_grade_class': selected_grade_class,
        'selected_term': selected_term,
        'grade_choices': GRADE_CHOICES,
        'assessments': [('class_test', 'Class Test'), ('monthly_exam', 'Monthly Exam'), ('midterm_exam', 'Midterm Exam'), ('appraisal_exam', 'Appraisal Exam'), ('end_of_term_exam', 'End of Term Exam')],
        'terms': Term.objects.all(),
    }

    return render(request, 'performance.html', context)






import csv
import io
import re
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import StudentCSVUploadForm
from .models import Student
from django.contrib.auth.models import User
from django.db import IntegrityError

def upload_students(request):
    if request.method == "POST":
        form = StudentCSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES["csv_file"]
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'This is not a CSV file')
                return redirect("upload_students")

            try:
                decoded_file = csv_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                next(io_string)  # Skip header

                for row in csv.reader(io_string):
                    if len(row) < 6:
                        messages.warning(request, f"Skipping invalid row: {row}")
                        continue

                    name, grade, assessment_choice, parent_name, parent_email, parent_phone = row

                    name = name.strip()
                    if not name:
                        messages.warning(request, "Missing student name. Skipping row.")
                        continue

                    # Clean and format username
                    base_username = re.sub(r'\W+', '', name.lower())[:30]
                    if not base_username:
                        messages.warning(request, f"Invalid username for student '{name}'. Skipping.")
                        continue

                    username = base_username
                    counter = 1
                    while User.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1

                    try:
                        # Create user
                        user = User.objects.create_user(username=username)
                        user.email = parent_email or ''
                        user.set_password('defaultpass123')
                        user.save()

                        # Create student
                        Student.objects.create(
                            user=user,
                            name=name,
                            grade=grade,
                            assessment_choice=assessment_choice,
                            parent_name=parent_name or None,
                            parent_email=parent_email or None,
                            parent_phone=parent_phone or None
                        )
                    except IntegrityError:
                        messages.warning(request, f"Student '{name}' with phone {parent_phone} already exists. Skipping.")
                        continue

                messages.success(request, "Students uploaded successfully!")
                return redirect("upload_students")

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
                return redirect("upload_students")

    else:
        form = StudentCSVUploadForm()

    return render(request, 'upload_students.html', {'form': form})






from django.shortcuts import render
from django.db.models import Sum, Q
from django.http import HttpResponse
import csv
from .models import Student, AssessmentScore, Term

GRADE_CHOICES = [
    ('7a', '7A'), ('7b', '7B'), ('7c', '7C'),
    ('8a', '8A'), ('8b', '8B'),
    ('9', '9')
]

ASSESSMENT_CHOICES = [
    ('class_test', 'Class Test'),
    ('monthly_exam', 'Monthly Exam'),
    ('midterm_exam', 'Midterm Exam'),
    ('appraisal_exam', 'Appraisal Exam'),
    ('end_of_term_exam', 'End of Term Exam'),
]

def export_analysis_csv(request):
    selected_grade = request.GET.get("grade", "")
    selected_category = request.GET.get("category", "class_test")
    selected_term_id = request.GET.get("term", "")

    # Get the term name if exists
    term_name = "All Terms"
    if selected_term_id:
        try:
            term = Term.objects.get(id=selected_term_id)
            term_name = term.name
        except Term.DoesNotExist:
            term_name = "Unknown Term"

    students = Student.objects.all()

    # Filter students by grade
    if selected_grade:
        if selected_grade == "7":
            students = students.filter(grade__startswith="7")
        elif selected_grade == "8":
            students = students.filter(grade__startswith="8")
        else:
            students = students.filter(grade=selected_grade)

    # Build filter for assessment scores
    score_filter = Q(assessmentscore__assessment_type=selected_category)
    if selected_term_id:
        score_filter &= Q(assessmentscore__term_id=selected_term_id)

    # Annotate and sort students by score
    student_scores = students.annotate(
        total_score=Sum("assessmentscore__score", filter=score_filter)
    ).order_by("-total_score")

    # Prepare the CSV response
    filename = f"student_analysis_{selected_grade}_{selected_category}_{term_name}.csv".replace(" ", "_").lower()
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # School header and report title
    writer.writerow(["📘 SCHOOL NAME HERE"])
    writer.writerow(["Student Performance Analysis"])
    writer.writerow([f"Grade: {selected_grade.upper() or 'All'}", f"Assessment: {dict(ASSESSMENT_CHOICES).get(selected_category, selected_category)}", f"Term: {term_name}"])
    writer.writerow([])  # Blank line

    # Table headers
    writer.writerow(["Rank", "Student ID", "Student Name", "Grade", "Total Score"])

    # Write student scores
    for rank, student in enumerate(student_scores, start=1):
        writer.writerow([rank, student.student_id, student.name, student.grade.upper(), student.total_score or 0])

    writer.writerow([])  # Blank line
    writer.writerow(["Generated by Teacher Dashboard"])

    return response



from django.shortcuts import render, redirect
from .models import AcademicCalendar
from .forms import AcademicCalendarForm
from django.utils.timezone import now
from django.contrib import messages

def add_event(request):
    if request.method == 'POST':
        form = AcademicCalendarForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '📅 Event successfully added!')
            return redirect('add_event')  # Redirect to calendar page after submission
        else:
            messages.error(request, '❌ Error: There was a problem with the form submission. Please try again.')
    else:
        form = AcademicCalendarForm()

    return render(request, 'add_event.html', {'form': form})



from django.shortcuts import render, get_object_or_404, redirect
from .models import Teacher
from .forms import TeacherRegistrationForm
from django.contrib import messages


def teacher_list(request):
    teachers = Teacher.objects.all()  # Retrieve all teachers from the database
    return render(request, 'teacher_list.html', {'teachers': teachers})



def teacher_edit(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == "POST":
        form = TeacherRegistrationForm(request.POST, instance=teacher, hide_password=True)
        if form.is_valid():
            form.save()
            messages.success(request, "Teacher details updated successfully!")
            return redirect('teacher_list')
    else:
        form = TeacherRegistrationForm(instance=teacher, hide_password=True)

    return render(request, 'teacher_edit.html', {'form': form, 'teacher': teacher})


# View to delete a teacher
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == "POST":
        teacher.delete()
        messages.success(request, "Teacher deleted successfully!")
        return redirect('teacher_list')  # Redirect to the teacher list after successful deletion

    return render(request, 'teacher_delete.html', {'teacher': teacher})



def index(request):
 
    return render(request, 'index.html')


from django.contrib.auth import logout
from django.shortcuts import redirect

def logout_view(request):
    logout(request)
    return redirect('home')  # Redirect to the home page after logout






















